import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
from calendar import monthrange
import os
from pathlib import Path
from tkinter import Tk, filedialog, messagebox
import tkinter as tk
from tkinter import scrolledtext
import sys
from io import StringIO

class VDIUserFilterApp:
    def __init__(self, input_file, target_year=None, target_month=None):
        """Initialize the VDI User Filter App"""
        self.input_file = input_file
        self.df_input = None
        self.df_computation = None
        self.output_file = None
        self.target_year = target_year if target_year else datetime.now().year
        self.target_month = target_month if target_month else datetime.now().month
        
    def read_excel(self):
        """Read the input Excel file"""
        try:
            # Try to read with different sheet names
            try:
                self.df_input = pd.read_excel(self.input_file, sheet_name=0)
            except Exception:
                # If that fails, try to get available sheets
                xl_file = pd.ExcelFile(self.input_file)
                print(f"Available sheets: {xl_file.sheet_names}")
                self.df_input = pd.read_excel(self.input_file, sheet_name=xl_file.sheet_names[0])
            
            print(f"✓ Successfully read input file: {self.input_file}")
            print(f"  Total rows: {len(self.df_input)}")
            
            # Normalize column names by stripping whitespace
            self.df_input.columns = self.df_input.columns.str.strip()
            
            # Handle unnamed columns (overflow of associated display name)
            unnamed_cols = [col for col in self.df_input.columns if col.startswith('Unnamed')]
            if unnamed_cols:
                for idx, col in enumerate(unnamed_cols):
                    self.df_input.rename(columns={col: f'(overflow of associated display name)'}, inplace=True)
            
            print(f"  Normalized columns: {list(self.df_input.columns)}")
            
            return True
        except Exception as e:
            print(f"✗ Error reading Excel file: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def parse_datetime(self, datetime_str):
        """Parse datetime string in format: 12/18/2025 2:29:58 PM"""
        try:
            # Strip whitespace before parsing
            if pd.isna(datetime_str):
                return pd.NaT
            datetime_str = str(datetime_str).strip()
            if datetime_str == '':
                return pd.NaT
            return pd.to_datetime(datetime_str, format='%m/%d/%Y %I:%M:%S %p')
        except:
            return pd.NaT
    
    def format_datetime_str(self, dt):
        """Format datetime to: 12/18/2025 2:29:58 PM"""
        if pd.isna(dt):
            return ""
        return dt.strftime('%m/%d/%Y %I:%M:%S %p')
    
    def format_date_str(self, dt):
        """Format date to: 12/18/2025"""
        if pd.isna(dt):
            return ""
        return dt.strftime('%m/%d/%Y')
    
    def format_time_str(self, dt):
        """Format time to: 2:29:58 PM"""
        if pd.isna(dt):
            return ""
        return dt.strftime('%I:%M:%S %p')
    
    def get_last_day_of_month(self, dt):
        """Get the last day of the month from the target month/year"""
        if pd.isna(dt):
            return pd.NaT
        # Use the target month/year instead of the dt's month/year
        last_day = monthrange(self.target_year, self.target_month)[1]
        return datetime(self.target_year, self.target_month, last_day)
    
    def calculate_inactivity_time(self, end_time):
        """Calculate time difference from end_time to 11:59:59 PM in format: x hours, x minutes, x seconds"""
        if pd.isna(end_time):
            return ""
        
        # Check if end_time is a string with only whitespace
        if isinstance(end_time, str) and end_time.strip() == '':
            return ""
        
        end_dt = end_time
        eod_time = end_dt.replace(hour=23, minute=59, second=59, microsecond=0)
        
        if eod_time < end_dt:
            # If end_dt is already past 11:59:59 PM, add a day
            eod_time = eod_time + timedelta(days=1)
        
        delta = eod_time - end_dt
        
        hours = delta.seconds // 3600
        minutes = (delta.seconds % 3600) // 60
        seconds = delta.seconds % 60
        
        return f"{hours} hours, {minutes} minutes, {seconds} seconds"
    
    def calculate_inactivity_days(self, end_date, last_date_of_month):
        """Calculate the difference in days between end_date and last_date_of_month (inclusive)"""
        if pd.isna(end_date) or pd.isna(last_date_of_month):
            return 0
        delta = last_date_of_month - end_date
        # Add 1 to include both the start and end dates in the count
        return delta.days + 1
    
    def generate_computation_worksheet(self):
        """Generate the Computation worksheet with all calculations"""
        df_comp = self.df_input.copy()
        
        # Strip whitespace from Session Start Time and Session End Time
        df_comp['Session Start Time'] = df_comp['Session Start Time'].astype(str).str.strip()
        df_comp['Session End Time'] = df_comp['Session End Time'].astype(str).str.strip()
        
        # Parse datetime columns
        df_comp['Session Start Time Parsed'] = df_comp['Session Start Time'].apply(self.parse_datetime)
        df_comp['Session End Time Parsed'] = df_comp['Session End Time'].apply(self.parse_datetime)
        
        # Calculate computed columns
        df_comp['End Date'] = df_comp['Session End Time Parsed'].apply(self.format_date_str)
        df_comp['End Time'] = df_comp['Session End Time Parsed'].apply(self.format_time_str)
        df_comp['Last Date of Month'] = df_comp['Session Start Time Parsed'].apply(
            lambda x: self.format_date_str(self.get_last_day_of_month(x))
        )
        df_comp['(Time)'] = '11:59:59 PM'
        
        # Calculate Length of Inactivity
        df_comp['Length of Inactivity'] = df_comp.apply(
            lambda row: f"{self.calculate_inactivity_days(row['Session End Time Parsed'], self.get_last_day_of_month(row['Session Start Time Parsed']))} days",
            axis=1
        )
        
        # Calculate Length of Inactivity Time
        df_comp['(Length of Inactivity Time)'] = df_comp['Session End Time Parsed'].apply(
            self.calculate_inactivity_time
        )
        
        # Reorder columns for output
        columns_order = [
            'Associated User',
            'Session Start Time',
            'Session End Time',
            'Endpoint Name',
            'Machine Name',
            'Delivery Group Name',
            'IP Address',
            'Associated User Display Name',
            'End Date',
            'End Time',
            'Last Date of Month',
            '(Time)',
            'Length of Inactivity',
            '(Length of Inactivity Time)'
        ]
        
        # Select only the final columns (drop parsed columns)
        df_comp = df_comp[columns_order]
        
        self.df_computation = df_comp
        print("✓ Computation worksheet generated")
        return df_comp
    
    def get_most_recent_sessions(self):
        """Get the most recent session for each user (case sensitive)"""
        df_with_parsed = self.df_computation.copy()
        
        # Parse datetime for sorting by most recent session end time
        df_with_parsed['Session End Time Parsed'] = df_with_parsed['Session End Time'].apply(self.parse_datetime)
        
        # Sort by Associated User and Session End Time, keep only the most recent
        df_recent = df_with_parsed.sort_values('Session End Time Parsed', ascending=False).drop_duplicates(
            subset=['Associated User'], keep='first'
        ).sort_index()
        
        return df_recent
    
    def create_summary_worksheet(self):
        """Create the SUMMARY worksheet"""
        df_recent = self.get_most_recent_sessions()
        
        # Calculate metrics
        total_sessions = len(self.df_computation)
        unique_users = self.df_computation['Associated User'].nunique()
        
        # Users with inactivity >= 20 days
        df_recent['Inactivity Days'] = df_recent['Length of Inactivity'].str.extract(r'(\d+)').astype(int)
        inactive_users_count = len(df_recent[df_recent['Inactivity Days'] >= 20])
        
        sessions_without_user = len(self.df_computation[
            self.df_computation['Associated User'].isna() | 
            (self.df_computation['Associated User'].astype(str).str.strip() == '')
        ])
        sessions_without_ip = len(self.df_computation[
            self.df_computation['IP Address'].isna() | 
            (self.df_computation['IP Address'].astype(str).str.strip() == '')
        ])
        sessions_without_display_name = len(self.df_computation[
            self.df_computation['Associated User Display Name'].isna() | 
            (self.df_computation['Associated User Display Name'].astype(str).str.strip() == '')
        ])
        
        # Use the target month and year
        month_name = datetime(self.target_year, self.target_month, 1).strftime('%B')
        year = self.target_year
        
        # Create summary dataframe
        summary_data = {
            'Metric': [
                f'VDI User Session Report {month_name} {year}',
                '',
                'Total VDI Session',
                'Total VDI User',
                'User with Last Session 20 Days Ago',
                'Session w/o Associated User',
                'Session w/o IP Address',
                'Session w/o Associated User Display Name'
            ],
            'Value': [
                '',
                '',
                total_sessions,
                unique_users,
                inactive_users_count,
                sessions_without_user,
                sessions_without_ip,
                sessions_without_display_name
            ]
        }
        
        df_summary = pd.DataFrame(summary_data)
        print("✓ Summary worksheet created")
        return df_summary
    
    def create_inactive_users_worksheet(self):
        """Create the Inactive Users worksheet (blank for now)"""
        df_inactive = pd.DataFrame()
        print("✓ Inactive Users worksheet created")
        return df_inactive
    
    def create_over_20_days_worksheet(self):
        """Create the >20 Days Ago User Session worksheet"""
        df_recent = self.get_most_recent_sessions()
        
        # Filter for inactivity > 20 days and time > 0 seconds
        df_recent['Inactivity Days'] = df_recent['Length of Inactivity'].str.extract(r'(\d+)').astype(int)
        
        # Extract total seconds from (Length of Inactivity Time)
        def time_to_seconds(time_str):
            if pd.isna(time_str) or time_str == '':
                return 0
            parts = time_str.split(',')
            hours = int(parts[0].split()[0]) if len(parts) > 0 else 0
            minutes = int(parts[1].split()[0]) if len(parts) > 1 else 0
            seconds = int(parts[2].split()[0]) if len(parts) > 2 else 0
            return hours * 3600 + minutes * 60 + seconds
        
        df_recent['Inactivity Seconds'] = df_recent['(Length of Inactivity Time)'].apply(time_to_seconds)
        
        df_over_20 = df_recent[(df_recent['Inactivity Days'] >= 20) & (df_recent['Inactivity Seconds'] > 0)].copy()

        # Sort alphabetically by Associated User
        df_over_20 = df_over_20.sort_values('Associated User')
        
        # Keep only the computation columns
        columns_order = [
            'Associated User',
            'Session Start Time',
            'Session End Time',
            'Endpoint Name',
            'Machine Name',
            'Delivery Group Name',
            'IP Address',
            'Associated User Display Name',
            'End Date',
            'End Time',
            'Last Date of Month',
            '(Time)',
            'Length of Inactivity',
            '(Length of Inactivity Time)'
        ]
        
        df_over_20 = df_over_20[[col for col in columns_order if col in df_over_20.columns]]
        
        print("✓ >20 Days Ago User Session worksheet created")
        return df_over_20
    
    def create_list_of_users_worksheet(self):
        """Create the List of Users worksheet"""
        unique_users = self.df_computation['Associated User'].dropna().unique()
        unique_users = sorted(unique_users)  # Case sensitive sorting
        
        df_users = pd.DataFrame({
            'Associated User': unique_users
        })
        
        print("✓ List of Users worksheet created")
        return df_users
    
    def create_without_associated_user_worksheet(self):
        """Create the without Associated User worksheet"""
        try:
            if 'Associated User' not in self.df_input.columns:
                print(f"  Warning: 'Associated User' column not found. Available columns: {list(self.df_input.columns)}")
                return pd.DataFrame()
            # Treat whitespace-only cells as empty
            df_without_user = self.df_input[
                self.df_input['Associated User'].isna() | 
                (self.df_input['Associated User'].astype(str).str.strip() == '')
            ].copy()
            print(f"✓ Without Associated User worksheet created ({len(df_without_user)} rows)")
            return df_without_user
        except Exception as e:
            print(f"✗ Error creating without Associated User worksheet: {e}")
            return pd.DataFrame()
    
    def create_without_ip_address_worksheet(self):
        """Create the without IP Address worksheet"""
        try:
            if 'IP Address' not in self.df_input.columns:
                print(f"  Warning: 'IP Address' column not found. Available columns: {list(self.df_input.columns)}")
                return pd.DataFrame()
            # Treat whitespace-only cells as empty
            df_without_ip = self.df_input[
                self.df_input['IP Address'].isna() | 
                (self.df_input['IP Address'].astype(str).str.strip() == '')
            ].copy()
            print(f"✓ Without IP Address worksheet created ({len(df_without_ip)} rows)")
            return df_without_ip
        except Exception as e:
            print(f"✗ Error creating without IP Address worksheet: {e}")
            return pd.DataFrame()
    
    def create_without_display_name_worksheet(self):
        """Create the without User Display Name worksheet"""
        try:
            if 'Associated User Display Name' not in self.df_input.columns:
                print(f"  Warning: 'Associated User Display Name' column not found. Available columns: {list(self.df_input.columns)}")
                return pd.DataFrame()
            # Treat whitespace-only cells as empty
            df_without_display = self.df_input[
                self.df_input['Associated User Display Name'].isna() | 
                (self.df_input['Associated User Display Name'].astype(str).str.strip() == '')
            ].copy()
            print(f"✓ Without User Display Name worksheet created ({len(df_without_display)} rows)")
            return df_without_display
        except Exception as e:
            print(f"✗ Error creating without User Display Name worksheet: {e}")
            return pd.DataFrame()
    
    def apply_worksheet_colors(self, workbook):
        """Apply colors to worksheet tabs"""
        # Red for Summary
        workbook['SUMMARY'].sheet_properties.tabColor = "FF0000"
        
        # Green for Computation
        workbook['Computation'].sheet_properties.tabColor = "00B050"
        
        # Orange for Inactive Users
        workbook['Inactive Users'].sheet_properties.tabColor = "FFC000"
        
        print("✓ Worksheet colors applied")
    
    def autofit_columns(self, workbook):
        """Autofit all columns in all worksheets"""
        from openpyxl.utils import get_column_letter
        
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            
            # Iterate through all columns
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                # Find the maximum length of content in this column
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # Set the column width with some padding
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long content
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print("✓ Column widths auto-fitted")
    
    def save_to_excel(self, output_file):
        """Save all worksheets to Excel file"""
        try:
            # Close any existing file if it's open
            if os.path.exists(output_file):
                try:
                    os.remove(output_file)
                    print(f"  Removed existing output file")
                except PermissionError:
                    print(f"  Warning: Could not remove existing file (may be open), will attempt to overwrite...")
            
            # Create Excel writer
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Write Summary
                df_summary = self.create_summary_worksheet()
                df_summary.to_excel(writer, sheet_name='SUMMARY', index=False)
                
                # Write Computation
                self.df_computation.to_excel(writer, sheet_name='Computation', index=False)
                
                # Write Inactive Users (blank)
                df_inactive = self.create_inactive_users_worksheet()
                if len(df_inactive) > 0:
                    df_inactive.to_excel(writer, sheet_name='Inactive Users', index=False)
                else:
                    empty_df = pd.DataFrame()
                    empty_df.to_excel(writer, sheet_name='Inactive Users', index=False)
                
                # Write >20 Days Ago User Session
                df_over_20 = self.create_over_20_days_worksheet()
                if len(df_over_20) > 0:
                    df_over_20.to_excel(writer, sheet_name='>20 Days Ago User Session', index=False)
                else:
                    empty_df = pd.DataFrame()
                    empty_df.to_excel(writer, sheet_name='>20 Days Ago User Session', index=False)
                
                # Write List of Users
                df_users = self.create_list_of_users_worksheet()
                df_users.to_excel(writer, sheet_name='List of Users', index=False)
                
                # Write without Associated User
                df_without_user = self.create_without_associated_user_worksheet()
                if len(df_without_user) > 0:
                    df_without_user.to_excel(writer, sheet_name='without Associated User', index=False)
                else:
                    empty_df = pd.DataFrame()
                    empty_df.to_excel(writer, sheet_name='without Associated User', index=False)
                
                # Write without IP Address
                df_without_ip = self.create_without_ip_address_worksheet()
                if len(df_without_ip) > 0:
                    df_without_ip.to_excel(writer, sheet_name='without IP address', index=False)
                else:
                    empty_df = pd.DataFrame()
                    empty_df.to_excel(writer, sheet_name='without IP address', index=False)
                
                # Write without User Display Name
                df_without_display = self.create_without_display_name_worksheet()
                if len(df_without_display) > 0:
                    df_without_display.to_excel(writer, sheet_name='without User Display Name', index=False)
                else:
                    empty_df = pd.DataFrame()
                    empty_df.to_excel(writer, sheet_name='without User Display Name', index=False)
            
            # Apply colors
            workbook = openpyxl.load_workbook(output_file)
            
            # Ensure all sheets are visible
            for sheet in workbook.sheetnames:
                workbook[sheet].sheet_state = 'visible'
            
            self.apply_worksheet_colors(workbook)
            self.autofit_columns(workbook)
            workbook.save(output_file)
            workbook.close()
            
            print(f"✓ Excel file saved successfully: {output_file}")
            return True
            
        except PermissionError as e:
            print(f"✗ Permission denied saving Excel file: {output_file}")
            print(f"  The file may be open in another program. Please close it and try again.")
            return False
        except Exception as e:
            print(f"✗ Error saving Excel file: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def process(self, output_file):
        """Main processing function"""
        print("\n" + "="*60)
        print("VDI User Filter App - Processing Started")
        print("="*60 + "\n")
        
        try:
            # Step 1: Read input file
            if not self.read_excel():
                return False
            
            # Step 2: Generate Computation worksheet
            self.generate_computation_worksheet()
            
            # Step 3: Save to Excel
            if not self.save_to_excel(output_file):
                return False
            
            print("\n" + "="*60)
            print("Processing Completed Successfully!")
            print("="*60 + "\n")
            return True
        
        except Exception as e:
            print(f"\n✗ ERROR: {str(e)}")
            print(f"Error type: {type(e).__name__}")
            import traceback
            print("\nFull traceback:")
            traceback.print_exc()
            return False


def get_month_year_dialog():
    """Show dialog to get target month and year from user"""
    dialog_window = tk.Tk()
    dialog_window.title("Select Report Month and Year")
    dialog_window.geometry("420x260")
    dialog_window.minsize(360, 220)
    dialog_window.resizable(True, True)
    dialog_window.configure(padx=16, pady=12)
    
    result = {'month': datetime.now().month, 'year': datetime.now().year}
    
    # Layout: two-column grid with consistent spacing
    dialog_window.grid_columnconfigure(0, weight=0)
    dialog_window.grid_columnconfigure(1, weight=1)
    dialog_window.grid_rowconfigure(2, weight=1)

    tk.Label(
        dialog_window,
        text="Choose the report period",
        font=("Arial", 11, "bold")
    ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))

    # Month selection
    tk.Label(dialog_window, text="Report Month", font=("Arial", 10)).grid(
        row=1, column=0, sticky="w", padx=(0, 8), pady=(0, 8)
    )
    months = ['January', 'February', 'March', 'April', 'May', 'June',
              'July', 'August', 'September', 'October', 'November', 'December']
    month_var = tk.StringVar(value=months[datetime.now().month - 1])
    month_dropdown = tk.OptionMenu(dialog_window, month_var, *months)
    month_dropdown.config(width=24)
    month_dropdown.grid(row=1, column=1, sticky="ew", pady=(0, 8))
    
    # Year selection
    tk.Label(dialog_window, text="Report Year", font=("Arial", 10)).grid(
        row=2, column=0, sticky="w", padx=(0, 8), pady=(0, 8)
    )
    year_var = tk.IntVar(value=datetime.now().year)
    year_spinbox = tk.Spinbox(
        dialog_window,
        from_=2000,
        to=2100,
        textvariable=year_var,
        width=10,
        font=("Arial", 10)
    )
    year_spinbox.grid(row=2, column=1, sticky="w", pady=(0, 8))
    
    def on_ok():
        result['month'] = months.index(month_var.get()) + 1
        result['year'] = year_var.get()
        dialog_window.destroy()
    
    def on_cancel():
        dialog_window.destroy()
    
    # Buttons
    button_frame = tk.Frame(dialog_window)
    button_frame.grid(row=3, column=0, columnspan=2, sticky="e", pady=(8, 0))
    cancel_button = tk.Button(button_frame, text="Cancel", command=on_cancel, width=12)
    ok_button = tk.Button(button_frame, text="OK", command=on_ok, width=12)
    cancel_button.pack(side=tk.RIGHT, padx=(6, 0))
    ok_button.pack(side=tk.RIGHT)

    # Accessibility and usability enhancements
    month_dropdown.focus_set()
    dialog_window.bind("<Return>", lambda event: on_ok())
    dialog_window.bind("<Escape>", lambda event: on_cancel())
    
    dialog_window.mainloop()
    
    return result['month'], result['year']


def main():
    """Main entry point with file picker and log display"""
    # Create hidden root window for file picker
    root = Tk()
    root.withdraw()
    
    # Open file dialog
    input_file = filedialog.askopenfilename(
        title="Select Input Excel File",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    
    # Check if user cancelled
    if not input_file:
        messagebox.showwarning("Cancelled", "No file selected. Application will exit.")
        root.destroy()
        return
    
    # Check if file exists
    if not os.path.exists(input_file):
        messagebox.showerror("Error", f"File not found: {input_file}")
        root.destroy()
        return
    
    root.destroy()
    
    # Ask user for target month and year
    target_month, target_year = get_month_year_dialog()
    
    # Generate output file name using selected month and year
    input_dir = os.path.dirname(input_file)
    month_name = datetime(target_year, target_month, 1).strftime('%B')
    output_file = os.path.join(input_dir, f"VDI_User_Filter_Report_{month_name}_{target_year}.xlsx")
    
    # Create and run the app with target month/year
    app = VDIUserFilterApp(input_file, target_year=target_year, target_month=target_month)
    success = app.process(output_file)
    
    # Show completion message in popup
    if success:
        messagebox.showinfo(
            "Success",
            f"Report generated successfully!\n\nOutput file:\n{output_file}"
        )
    else:
        messagebox.showerror(
            "Error",
            f"Failed to generate report.\n\nPlease check the console output for more information."
        )

if __name__ == "__main__":
    main()
